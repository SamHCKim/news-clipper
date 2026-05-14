"""Core logic for News Clipper — pure functions reused by both the CLI (news.py)
and the Streamlit UI (app.py). No console I/O or argparse here.
"""

from __future__ import annotations

import difflib
import re
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import IO, Union
from urllib.parse import quote, urlparse, urlunparse

import feedparser
from openpyxl import Workbook

LOCALES = [
    # (hl, gl, lang_for_ceid)  -> Google News RSS keeps articles in their source language
    ("ko", "KR", "ko"),
    ("en-US", "US", "en"),
]

COLUMNS = ["매칭 키워드", "제목", "링크", "언론사", "발행일시"]
COLUMNS_MULTI = ["검색 쿼리", "매칭 키워드", "제목", "링크", "언론사", "발행일시"]
FUZZY_TITLE_THRESHOLD = 0.85
MIN_TITLE_LEN_FUZZY = 10

_OPERATOR_TOKENS = {"AND", "OR", "NOT"}


def build_rss_url(query: str, days: int, hl: str, gl: str, lang: str) -> str:
    q = f"{query} when:{days}d"
    ceid = f"{gl}:{lang}"
    return (
        "https://news.google.com/rss/search"
        f"?q={quote(q)}&hl={hl}&gl={gl}&ceid={ceid}"
    )


def normalize_url(url: str) -> str:
    try:
        p = urlparse(url)
        return urlunparse((p.scheme, p.netloc, p.path, "", "", ""))
    except Exception:
        return url


def extract_source(entry) -> str:
    src = getattr(entry, "source", None)
    if src is not None:
        title = getattr(src, "title", None) or (
            src.get("title") if isinstance(src, dict) else None
        )
        if title:
            return title
    title = entry.get("title", "")
    if " - " in title:
        return title.rsplit(" - ", 1)[1]
    return ""


def strip_source_from_title(title: str, source: str) -> str:
    if source and title.endswith(f" - {source}"):
        return title[: -(len(source) + 3)]
    return title


def fetch_news(query: str, days: int) -> list[dict]:
    items: list[dict] = []
    for hl, gl, lang in LOCALES:
        url = build_rss_url(query, days, hl, gl, lang)
        feed = feedparser.parse(url)
        for e in feed.entries:
            source = extract_source(e)
            title = strip_source_from_title(e.get("title", ""), source)
            link = e.get("link", "")
            items.append(
                {
                    "title": title,
                    "link": link,
                    "source": source,
                    "published": e.get("published", ""),
                    "norm_link": normalize_url(link),
                    "locale": f"{hl}/{gl}",
                }
            )
    return items


def dedupe(items: list[dict]) -> list[dict]:
    seen: dict[str, dict] = {}
    for it in items:
        key = it["norm_link"] or it["title"]
        if key and key not in seen:
            seen[key] = it
    return list(seen.values())


def load_whitelist(value: str | None) -> list[str]:
    if not value:
        return []
    p = Path(value)
    if p.exists():
        return [
            line.strip()
            for line in p.read_text(encoding="utf-8").splitlines()
            if line.strip() and not line.startswith("#")
        ]
    return [s.strip() for s in value.split(",") if s.strip()]


def filter_trusted(items: list[dict], whitelist: list[str]) -> list[dict]:
    if not whitelist:
        return items
    wl = [w.lower() for w in whitelist]

    def matches(it: dict) -> bool:
        haystack = (it["source"] + " " + it["link"]).lower()
        return any(w in haystack for w in wl)

    return [it for it in items if matches(it)]


def find_matched_keywords(title: str, query: str) -> list[str]:
    """Return list of bare tokens from the query that appear in the title.

    Ignores Google search operators (AND/OR/NOT), negations (-foo) and
    field selectors (site:...).
    """
    tokens = re.findall(r'"([^"]+)"|([^\s()]+)', query)
    candidates: list[str] = []
    for quoted, plain in tokens:
        tok = quoted or plain
        if not tok:
            continue
        if tok.upper() in _OPERATOR_TOKENS:
            continue
        if tok.startswith("-") or ":" in tok:
            continue
        candidates.append(tok)
    lower_title = title.lower()
    return [t for t in candidates if t.lower() in lower_title]


def rows_for(items: list[dict], query: str) -> list[list[str]]:
    """Single-query row layout (CLI output)."""
    return [
        [
            ", ".join(find_matched_keywords(it["title"], query)),
            it["title"],
            it["link"],
            it["source"],
            it["published"],
        ]
        for it in items
    ]


def write_xlsx(items: list[dict], query: str, out_path: Union[str, IO[bytes]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "News"
    ws.append(COLUMNS)
    for row in rows_for(items, query):
        ws.append(row)
    widths = [20, 60, 60, 20, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(out_path)


def write_gsheet(items: list[dict], query: str, sheet_id: str, credentials_path: str) -> None:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        sys.exit(
            "gspread and google-auth are required for --output gsheet. "
            "Install with: pip install gspread google-auth"
        )
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    ws = sh.sheet1
    if not ws.get_all_values():
        ws.append_row(COLUMNS)
    rows = rows_for(items, query)
    if rows:
        ws.append_rows(rows)


# ----------------------------------------------------------------------
# Multi-query support (used by the Streamlit UI)
# ----------------------------------------------------------------------


def fetch_for_queries(
    queries: list[str], days: int, max_workers: int = 7
) -> list[dict]:
    """Run fetch_news() in parallel for each query, tagging each item with the
    originating query string. Returns a flat list (no dedup yet)."""
    queries = [q for q in queries if q and q.strip()]
    if not queries:
        return []

    def _one(q: str) -> list[dict]:
        results = fetch_news(q, days)
        for r in results:
            r["matched_query"] = q
        return results

    out: list[dict] = []
    with ThreadPoolExecutor(max_workers=max(1, min(len(queries), max_workers))) as ex:
        for batch in ex.map(_one, queries):
            out.extend(batch)
    return out


def dedupe_multi(items: list[dict]) -> list[dict]:
    """Dedup by normalized URL (exact) then by fuzzy title similarity.

    Pass 1 — exact key dedup: group by norm_link (or title as fallback),
    accumulating matched_queries across duplicates.

    Pass 2 — fuzzy title dedup: for every pair of surviving items whose titles
    are both at least MIN_TITLE_LEN_FUZZY characters, compute
    SequenceMatcher ratio; if >= FUZZY_TITLE_THRESHOLD treat them as the same
    article, keep the first-seen item, and union the matched_queries sets.
    O(n²) — acceptable for the typical ≤700-item result set.
    """
    # --- Pass 1: exact key dedup ---
    seen: dict[str, dict] = {}
    for it in items:
        key = it.get("norm_link") or it.get("title")
        if not key:
            continue
        if key not in seen:
            base = dict(it)
            base["matched_queries"] = {it.get("matched_query", "")} if it.get("matched_query") else set()
            seen[key] = base
        else:
            mq = it.get("matched_query")
            if mq:
                seen[key]["matched_queries"].add(mq)
    unique = list(seen.values())

    # --- Pass 2: fuzzy title dedup ---
    kept: list[dict] = []
    for candidate in unique:
        t_cand = (candidate.get("title") or "").strip()
        merged = False
        if len(t_cand) >= MIN_TITLE_LEN_FUZZY:
            for existing in kept:
                t_exist = (existing.get("title") or "").strip()
                if len(t_exist) >= MIN_TITLE_LEN_FUZZY:
                    sm = difflib.SequenceMatcher(None, t_exist, t_cand)
                    if sm.quick_ratio() < FUZZY_TITLE_THRESHOLD:
                        continue
                    if sm.ratio() >= FUZZY_TITLE_THRESHOLD:
                        existing["matched_queries"] |= candidate.get("matched_queries", set())
                        merged = True
                        break
        if not merged:
            kept.append(candidate)
    return kept


def rows_for_multi(items: list[dict]) -> list[list[str]]:
    """Multi-query row layout: [검색 쿼리, 매칭 키워드, 제목, 링크, 언론사, 발행일시]."""
    out: list[list[str]] = []
    for it in items:
        queries = sorted(q for q in it.get("matched_queries", []) if q)
        matched_tokens: set[str] = set()
        for q in queries:
            for tok in find_matched_keywords(it["title"], q):
                matched_tokens.add(tok)
        out.append(
            [
                ", ".join(queries),
                ", ".join(sorted(matched_tokens, key=str.lower)),
                it["title"],
                it["link"],
                it["source"],
                it["published"],
            ]
        )
    return out


def write_xlsx_multi(items: list[dict], out_path_or_buffer: Union[str, IO[bytes]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "News"
    ws.append(COLUMNS_MULTI)
    for row in rows_for_multi(items):
        ws.append(row)
    widths = [28, 20, 60, 60, 20, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(out_path_or_buffer)
